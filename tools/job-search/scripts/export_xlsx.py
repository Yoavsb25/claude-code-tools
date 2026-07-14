#!/usr/bin/env python3
"""
Optional Excel export for job-search shortlists. Requires `openpyxl` (not stdlib) — this is a
deliberate exception to job_tool.py's stdlib-only rule, kept in its own script so the core
search/profile/tracker bookkeeping never gains a hard third-party dependency.

Usage:
  python3 export_xlsx.py --input payload.json [--output path.xlsx]
  cat payload.json | python3 export_xlsx.py [--output path.xlsx]

Input JSON shape:
{
  "date": "2026-07-14",
  "sheets": [
    {
      "name": "Public Companies",          # sheet name, <=31 chars
      "title": "Public Companies — ...",   # title row shown above the table
      "header_color": "1F4E78",            # optional hex, no leading '#'
      "rows": [
        {
          "company": "Acme Corp", "status": "NASDAQ: ACME", "role": "Staff Backend Engineer",
          "location": "London", "skills_fit": 8, "why": "one-line rationale",
          "connections": ["Jane Doe", "John Smith"],   # optional, from `network match`
          "link": "https://..."
        }
      ]
    }
  ]
}

If `openpyxl` isn't installed, this prints one line to stderr with the install command and exits
1 — degrade by falling back to the markdown tables already in the conversation, never block the
rest of a job-search run on this being unavailable.
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "error: openpyxl not installed. Set up a venv and install it, e.g.:\n"
        "  python3 -m venv ~/.claude/skills/job-search/.venv\n"
        "  source ~/.claude/skills/job-search/.venv/bin/activate\n"
        "  pip install openpyxl\n"
        "then re-run this script with that venv's python3.",
        file=sys.stderr,
    )
    sys.exit(1)

COLUMNS = ["#", "Company", "Ticker / Status", "Role", "Location", "Skills Fit", "Why", "Connections", "Link"]
COL_WIDTHS = [4, 26, 16, 42, 22, 10, 46, 30, 14]
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def connections_text(names):
    return ", ".join(names) if names else "—"


def write_sheet(ws, sheet):
    title = sheet.get("title") or sheet["name"]
    header_color = sheet.get("header_color") or "1F4E78"
    rows = sheet.get("rows") or []

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    header_row = 3
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for i, row in enumerate(rows):
        r = header_row + 1 + i
        conn_text = connections_text(row.get("connections"))
        why = row.get("why") or ""
        values = [
            i + 1, row.get("company"), row.get("status"), row.get("role"),
            row.get("location"), row.get("skills_fit"), why, conn_text,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = WRAP if col_idx in (7, 8) else Alignment(vertical="top")
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="top")
        link = row.get("link")
        link_cell = ws.cell(row=r, column=9, value="Apply →" if link else "—")
        if link:
            link_cell.hyperlink = link
            link_cell.font = LINK_FONT
        link_cell.border = BORDER
        link_cell.alignment = Alignment(horizontal="center", vertical="top")

        conn_lines = max(1, (len(conn_text) // 30) + 1)
        why_lines = max(1, (len(why) // 40) + 1)
        ws.row_dimensions[r].height = 15 * max(2, conn_lines, why_lines)

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.row_dimensions[header_row].height = 20


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", help="Path to payload JSON (default: read from stdin)")
    parser.add_argument("--output", help="Output .xlsx path (default: ~/Desktop/Job-Search/searches/<date>-job-search-results.xlsx)")
    args = parser.parse_args()

    raw = Path(args.input).read_text(encoding="utf-8") if args.input else sys.stdin.read()
    payload = json.loads(raw)

    out_date = payload.get("date") or date.today().isoformat()
    output_path = Path(args.output) if args.output else Path(
        f"~/Desktop/Job-Search/searches/{out_date}-job-search-results.xlsx"
    ).expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    sheets = payload.get("sheets") or []
    if not sheets:
        print("error: payload has no 'sheets'", file=sys.stderr)
        sys.exit(1)

    ws = wb.active
    ws.title = sheets[0]["name"][:31]
    write_sheet(ws, sheets[0])
    for sheet in sheets[1:]:
        ws = wb.create_sheet(sheet["name"][:31])
        write_sheet(ws, sheet)

    wb.save(output_path)
    print(json.dumps({"output_path": str(output_path)}))


if __name__ == "__main__":
    main()
